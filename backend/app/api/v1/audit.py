"""Audit log endpoints: list, detail, Excel export."""

import io
from datetime import date, datetime, timezone

from fastapi import APIRouter, Depends, Query
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.responses import StreamingResponse

from app.api.deps import PermissionChecker
from app.database import get_session
from app.models.admin_user import AdminUser
from app.models.audit_log import AuditLog
from app.schemas.audit import AuditLogListResponse, AuditLogResponse

router = APIRouter(prefix="/audit", tags=["audit"])

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_log_response(log: AuditLog, admin_username: str | None) -> AuditLogResponse:
    return AuditLogResponse(
        id=log.id,
        admin_user_id=log.admin_user_id,
        admin_username=admin_username,
        ip_address=log.ip_address,
        user_agent=log.user_agent,
        action=log.action,
        module=log.module,
        resource_type=log.resource_type,
        resource_id=log.resource_id,
        before_data=log.before_data,
        after_data=log.after_data,
        description=log.description,
        created_at=log.created_at,
    )


async def _batch_admin_usernames(session: AsyncSession, logs: list[AuditLog]) -> dict[int, str]:
    """Batch-load admin usernames for a list of audit logs."""
    admin_ids = {log.admin_user_id for log in logs if log.admin_user_id}
    if not admin_ids:
        return {}
    result = await session.execute(
        select(AdminUser.id, AdminUser.username).where(AdminUser.id.in_(admin_ids))
    )
    return {row[0]: row[1] for row in result.all()}


def _parse_dates(
    start_date: str | None, end_date: str | None
) -> tuple[datetime | None, datetime | None]:
    start = None
    end = None
    if start_date:
        start = datetime.combine(
            datetime.strptime(start_date, "%Y-%m-%d").date(),
            datetime.min.time(),
            tzinfo=timezone.utc,
        )
    if end_date:
        end = datetime.combine(
            datetime.strptime(end_date, "%Y-%m-%d").date(),
            datetime.max.time(),
            tzinfo=timezone.utc,
        )
    return start, end


# ─── List Audit Logs ──────────────────────────────────────────────


@router.get("/logs", response_model=AuditLogListResponse)
async def list_audit_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    action: str | None = Query(None),
    module: str | None = Query(None),
    admin_user_id: int | None = Query(None),
    admin_username: str | None = Query(None, description="Admin username (partial match)"),
    start_date: str | None = Query(None, description="YYYY-MM-DD"),
    end_date: str | None = Query(None, description="YYYY-MM-DD"),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("audit_log.view")),
):
    base = select(AuditLog)

    if action:
        base = base.where(AuditLog.action == action)
    if module:
        base = base.where(AuditLog.module == module)
    if admin_user_id:
        base = base.where(AuditLog.admin_user_id == admin_user_id)
    if admin_username:
        safe_name = admin_username.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        admin_ids_subq = (
            select(AdminUser.id)
            .where(AdminUser.username.ilike(f"%{safe_name}%", escape="\\"))
            .scalar_subquery()
        )
        base = base.where(AuditLog.admin_user_id.in_(admin_ids_subq))

    start, end = _parse_dates(start_date, end_date)
    if start:
        base = base.where(AuditLog.created_at >= start)
    if end:
        base = base.where(AuditLog.created_at <= end)

    count_stmt = select(func.count()).select_from(base.subquery())
    total = (await session.execute(count_stmt)).scalar() or 0

    stmt = base.order_by(AuditLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await session.execute(stmt)
    logs = result.scalars().all()

    username_map = await _batch_admin_usernames(session, logs)
    items = [_build_log_response(log, username_map.get(log.admin_user_id)) for log in logs]
    return AuditLogListResponse(items=items, total=total, page=page, page_size=page_size)


# ─── Export Audit Logs (Excel) ── MUST be before {log_id} route ──


@router.get("/logs/export")
async def export_audit_logs(
    action: str | None = Query(None),
    module: str | None = Query(None),
    admin_user_id: int | None = Query(None),
    admin_username: str | None = Query(None, description="Admin username (partial match)"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("audit_log.export")),
):
    base = select(AuditLog)

    if action:
        base = base.where(AuditLog.action == action)
    if module:
        base = base.where(AuditLog.module == module)
    if admin_user_id:
        base = base.where(AuditLog.admin_user_id == admin_user_id)
    if admin_username:
        safe_name = admin_username.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        admin_ids_subq = (
            select(AdminUser.id)
            .where(AdminUser.username.ilike(f"%{safe_name}%", escape="\\"))
            .scalar_subquery()
        )
        base = base.where(AuditLog.admin_user_id.in_(admin_ids_subq))

    start, end = _parse_dates(start_date, end_date)
    if start:
        base = base.where(AuditLog.created_at >= start)
    if end:
        base = base.where(AuditLog.created_at <= end)

    stmt = base.order_by(AuditLog.created_at.desc()).limit(5000)
    result = await session.execute(stmt)
    logs = result.scalars().all()

    # Build Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    headers = [
        "ID",
        "Admin User ID",
        "Username",
        "Action",
        "Module",
        "Resource Type",
        "Resource ID",
        "IP Address",
        "Description",
        "Created At",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    username_map = await _batch_admin_usernames(session, logs)

    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=log.id)
        ws.cell(row=row_idx, column=2, value=log.admin_user_id)
        ws.cell(row=row_idx, column=3, value=username_map.get(log.admin_user_id, ""))
        ws.cell(row=row_idx, column=4, value=log.action)
        ws.cell(row=row_idx, column=5, value=log.module)
        ws.cell(row=row_idx, column=6, value=log.resource_type)
        ws.cell(row=row_idx, column=7, value=log.resource_id)
        ws.cell(row=row_idx, column=8, value=log.ip_address)
        ws.cell(row=row_idx, column=9, value=log.description)
        ws.cell(
            row=row_idx,
            column=10,
            value=log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
        )

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().strftime("%Y-%m-%d")
    filename = f"audit_logs_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type=EXCEL_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Get Audit Log Detail ────────────────────────────────────────


@router.get("/logs/{log_id}", response_model=AuditLogResponse)
async def get_audit_log(
    log_id: int,
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("audit_log.view")),
):
    from fastapi import HTTPException

    log = await session.get(AuditLog, log_id)
    if not log:
        raise HTTPException(status_code=404, detail="Audit log not found")
    username_map = await _batch_admin_usernames(session, [log])
    return _build_log_response(log, username_map.get(log.admin_user_id))
