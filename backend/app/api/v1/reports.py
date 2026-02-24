"""Report endpoints: agent, commission, financial reports with Excel export."""

import io
from datetime import date, datetime, time as time_type, timezone
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.responses import StreamingResponse

from app.api.deps import PermissionChecker
from app.database import get_session
from app.models.admin_user import AdminUser
from app.models.commission import CommissionLedger
from app.models.transaction import Transaction
from app.models.user import User
from app.schemas.report import (
    AgentReportItem,
    AgentReportResponse,
    CommissionReportByUser,
    CommissionReportItem,
    CommissionReportResponse,
    FinancialReportResponse,
)

router = APIRouter(prefix="/reports", tags=["reports"])

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _parse_dates(start_date: str | None, end_date: str | None) -> tuple[datetime, datetime]:
    today = datetime.now(timezone.utc).date()
    start = datetime.combine(
        datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else today.replace(day=1),
        time_type.min,
    )
    end = datetime.combine(
        datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else today,
        time_type.max,
    )
    return start, end


# ─── Agent Report ──────────────────────────────────────────────────

@router.get("/agents", response_model=AgentReportResponse)
async def agent_report(
    start_date: str | None = Query(None, description="YYYY-MM-DD"),
    end_date: str | None = Query(None, description="YYYY-MM-DD"),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("report.view")),
) -> AgentReportResponse:
    start, end = _parse_dates(start_date, end_date)

    agents_result = await session.execute(
        select(AdminUser).where(
            AdminUser.status == "active",
            AdminUser.role != "super_admin",
        )
    )
    agents = agents_result.scalars().all()
    agent_ids = [a.id for a in agents]

    # Batch: user count per agent via User.referrer_id
    user_count_stmt = (
        select(User.referrer_id, func.count().label("cnt"))
        .where(User.referrer_id.in_(agent_ids))
        .group_by(User.referrer_id)
    )
    user_count_result = await session.execute(user_count_stmt)
    user_count_map: dict[int, int] = {r.referrer_id: r.cnt for r in user_count_result.all()}

    # Batch: total bets per agent (rolling type, via recipient_user_id)
    bets_stmt = (
        select(
            CommissionLedger.recipient_user_id,
            func.coalesce(func.sum(CommissionLedger.source_amount), 0).label("total_bets"),
        )
        .where(
            CommissionLedger.recipient_user_id.in_(agent_ids),
            CommissionLedger.type == "rolling",
            CommissionLedger.created_at.between(start, end),
        )
        .group_by(CommissionLedger.recipient_user_id)
    )
    bets_result = await session.execute(bets_stmt)
    bets_map: dict[int, Decimal] = {r.recipient_user_id: Decimal(str(r.total_bets)) for r in bets_result.all()}

    # Batch: total commissions per agent (via recipient_user_id)
    comm_stmt = (
        select(
            CommissionLedger.recipient_user_id,
            func.coalesce(func.sum(CommissionLedger.commission_amount), 0).label("total_comm"),
        )
        .where(
            CommissionLedger.recipient_user_id.in_(agent_ids),
            CommissionLedger.created_at.between(start, end),
        )
        .group_by(CommissionLedger.recipient_user_id)
    )
    comm_result = await session.execute(comm_stmt)
    comm_map: dict[int, Decimal] = {r.recipient_user_id: Decimal(str(r.total_comm)) for r in comm_result.all()}

    items = [
        AgentReportItem(
            agent_id=agent.id,
            username=agent.username,
            agent_code=agent.agent_code,
            role=agent.role,
            total_users=user_count_map.get(agent.id, 0),
            total_bets=bets_map.get(agent.id, Decimal("0")),
            total_commissions=comm_map.get(agent.id, Decimal("0")),
        )
        for agent in agents
    ]

    return AgentReportResponse(
        items=items,
        start_date=start.strftime("%Y-%m-%d"),
        end_date=end.strftime("%Y-%m-%d"),
    )


# ─── Commission Report ────────────────────────────────────────────

@router.get("/commissions", response_model=CommissionReportResponse)
async def commission_report(
    start_date: str | None = Query(None, description="YYYY-MM-DD"),
    end_date: str | None = Query(None, description="YYYY-MM-DD"),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("report.view")),
) -> CommissionReportResponse:
    start, end = _parse_dates(start_date, end_date)

    # By type summary
    type_result = await session.execute(
        select(
            CommissionLedger.type,
            func.coalesce(func.sum(CommissionLedger.commission_amount), 0).label("total_amount"),
            func.count().label("count"),
        ).where(
            CommissionLedger.created_at.between(start, end),
        ).group_by(CommissionLedger.type)
    )
    items = [
        CommissionReportItem(type=r.type, total_amount=r.total_amount, count=r.count)
        for r in type_result.all()
    ]

    # By user breakdown (recipient_user_id → User)
    user_result = await session.execute(
        select(
            CommissionLedger.recipient_user_id,
            User.username,
            func.coalesce(func.sum(
                case((CommissionLedger.type == "rolling", CommissionLedger.commission_amount), else_=0)
            ), 0).label("rolling_total"),
            func.coalesce(func.sum(
                case((CommissionLedger.type == "losing", CommissionLedger.commission_amount), else_=0)
            ), 0).label("losing_total"),
        )
        .join(User, User.id == CommissionLedger.recipient_user_id)
        .where(CommissionLedger.created_at.between(start, end))
        .group_by(CommissionLedger.recipient_user_id, User.username)
    )
    by_user = [
        CommissionReportByUser(
            recipient_user_id=r.recipient_user_id,
            username=r.username,
            rolling_total=r.rolling_total,
            losing_total=r.losing_total,
        )
        for r in user_result.all()
    ]

    return CommissionReportResponse(
        items=items,
        by_user=by_user,
        start_date=start.strftime("%Y-%m-%d"),
        end_date=end.strftime("%Y-%m-%d"),
    )


# ─── Financial Report ─────────────────────────────────────────────

@router.get("/financial", response_model=FinancialReportResponse)
async def financial_report(
    start_date: str | None = Query(None, description="YYYY-MM-DD"),
    end_date: str | None = Query(None, description="YYYY-MM-DD"),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("report.view")),
) -> FinancialReportResponse:
    start, end = _parse_dates(start_date, end_date)

    # Deposits/withdrawals (approved only)
    tx_result = (await session.execute(
        select(
            func.coalesce(func.sum(
                case((Transaction.type == "deposit", Transaction.amount), else_=0)
            ), 0).label("deposits"),
            func.coalesce(func.sum(
                case((Transaction.type == "withdrawal", Transaction.amount), else_=0)
            ), 0).label("withdrawals"),
            func.coalesce(func.sum(
                case((Transaction.type == "deposit", 1), else_=0)
            ), 0).label("deposit_count"),
            func.coalesce(func.sum(
                case((Transaction.type == "withdrawal", 1), else_=0)
            ), 0).label("withdrawal_count"),
        ).where(
            Transaction.status == "approved",
            Transaction.created_at.between(start, end),
        )
    )).one()

    # Total commissions in period
    total_commissions = (await session.execute(
        select(func.coalesce(func.sum(CommissionLedger.commission_amount), 0)).where(
            CommissionLedger.created_at.between(start, end),
        )
    )).scalar() or 0

    deposits = Decimal(str(tx_result.deposits))
    withdrawals = Decimal(str(tx_result.withdrawals))

    return FinancialReportResponse(
        total_deposits=deposits,
        total_withdrawals=withdrawals,
        net_revenue=deposits - withdrawals,
        total_commissions=Decimal(str(total_commissions)),
        deposit_count=int(tx_result.deposit_count),
        withdrawal_count=int(tx_result.withdrawal_count),
        start_date=start.strftime("%Y-%m-%d"),
        end_date=end.strftime("%Y-%m-%d"),
    )


# ─── Excel Exports ────────────────────────────────────────────────

def _make_excel(headers: list[str], rows: list[list], sheet_name: str = "Report") -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/agents/export")
async def export_agent_report(
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("report.export")),
):
    report = await agent_report(start_date, end_date, session, current_user)
    headers = ["ID", "Username", "Agent Code", "Role", "Users", "Total Bets", "Total Commissions"]
    rows = [
        [i.agent_id, i.username, i.agent_code, i.role, i.total_users,
         float(i.total_bets), float(i.total_commissions)]
        for i in report.items
    ]
    buf = _make_excel(headers, rows, "Agent Report")
    filename = f"agent_report_{report.start_date}_{report.end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type=EXCEL_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/commissions/export")
async def export_commission_report(
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("report.export")),
):
    report = await commission_report(start_date, end_date, session, current_user)
    headers = ["User ID", "Username", "Rolling Total", "Losing Total"]
    rows = [
        [a.recipient_user_id, a.username, float(a.rolling_total), float(a.losing_total)]
        for a in report.by_user
    ]
    buf = _make_excel(headers, rows, "Commission Report")
    filename = f"commission_report_{report.start_date}_{report.end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type=EXCEL_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/financial/export")
async def export_financial_report(
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("report.export")),
):
    report = await financial_report(start_date, end_date, session, current_user)
    headers = ["Metric", "Value"]
    rows = [
        ["Total Deposits", float(report.total_deposits)],
        ["Total Withdrawals", float(report.total_withdrawals)],
        ["Net Revenue", float(report.net_revenue)],
        ["Total Commissions", float(report.total_commissions)],
        ["Deposit Count", report.deposit_count],
        ["Withdrawal Count", report.withdrawal_count],
        ["Period", f"{report.start_date} ~ {report.end_date}"],
    ]
    buf = _make_excel(headers, rows, "Financial Report")
    filename = f"financial_report_{report.start_date}_{report.end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type=EXCEL_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
